# -*- coding: utf-8 -*-
# Requires: pip install openpyxl
import re
import os
# 确保安装了兼容 Python 3 的 openpyxl 版本
# pip install openpyxl (推荐) 或 pip install openpyxl==2.6.4 (如果必须用旧版)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
#python -m venv myenv
#需要在虚拟环境下运行 myenv\Scripts\activate

# 定义需要提取数据的章节标题列表
TARGET_SECTIONS = [
    r'8\.1\.3\s+P3Y\s+CTVT\s+Cali\s+10A',
    r'8\.1\.6\s+P3Y\s+CTVT\s+Cali\s+1A',
    r'8\.4\s+50Hz\s+Verification2',
    r'8\.4\s+60Hz\s+Verification1',
    r'8\.4\s+60Hz\s+Verification2'
]

# 编译章节匹配的正则表达式
SECTION_PATTERNS = [re.compile(pattern) for pattern in TARGET_SECTIONS]


def sanitize_sheet_name(name, max_len=31):
    """Excel sheet names have limitations. Sanitize the filename for use as a sheet name."""
    # 移除或替换非法字符: \ / ? * [ ]
    illegal_chars = r'\/?*[]:'
    for char in illegal_chars:
        name = name.replace(char, '_')
    # 限制长度 (Excel 限制为 31 个字符)
    if len(name) > max_len:
        name = name[:max_len]
    # 移除首尾空格
    name = name.strip()
    # 确保名称不为空
    if not name:
        name = "Sheet"
    return name


def extract_data(filepath):
    """Extract Meas VAL and ANGLE data from file for specific sections."""
    all_section_data = []

    if not os.path.exists(filepath):
        print("  Error: File not found '%s'" % filepath)
        return all_section_data

    print("  Reading file: %s" % os.path.basename(filepath))

    current_section = None
    current_section_data = []
    in_target_section = False
    temp_channel_data = {}

    # Python 3: 默认使用 utf-8 编码打开文件，无需 decode
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            # Python 3: line 本身就是字符串，无需 decode
            line_u = line
            line_stripped = line_u.strip()

            # 检查是否进入新的目标章节
            for i, pattern in enumerate(SECTION_PATTERNS):
                if pattern.search(line_stripped):
                    # 如果之前在目标章节并且有数据，则保存之前的章节数据
                    if in_target_section and (current_section_data or temp_channel_data):
                        # 保存上一节未保存的临时数据（如果有的话）
                        if temp_channel_data: # 修正: 添加了条件判断
                            current_section_data.append(temp_channel_data)
                            temp_channel_data = {}
                        all_section_data.append({
                            'section_name': current_section,
                            'data': current_section_data
                        })

                    # 开始新的章节
                    current_section = TARGET_SECTIONS[i].replace(r'\s+', ' ').replace(r'\.', '.')
                    current_section_data = []
                    in_target_section = True
                    temp_channel_data = {}
                    # print("  Found target section: %s" % current_section) # 可选：打印找到的章节
                    break  # 找到匹配就跳出循环

            # 如果不在目标章节，继续读取下一行
            if not in_target_section:
                continue

            # 检查是否进入下一个类似 "数字.数字.数字" 或 "数字.数字" 开头的章节
            if re.match(r'^\d+\.\d+(?:\.\d+)?\s', line_stripped):
                # 保存当前章节数据
                if current_section_data or temp_channel_data: # 修正: 添加了条件判断
                    # 保存最后一组未保存的临时数据（如果有的话）
                    if temp_channel_data: # 修正: 添加了条件判断
                        current_section_data.append(temp_channel_data)
                        temp_channel_data = {}
                    all_section_data.append({
                        'section_name': current_section,
                        'data': current_section_data
                    })
                    # 重置状态
                    current_section = None
                    current_section_data = []
                    in_target_section = False
                continue

            # 在目标章节内，匹配 Meas VAL Check pattern
            val_match = re.search(
                r'Meas VAL Check<<C>> CH(\d+):\s*([-\d.]+)',
                line_u
            )
            if val_match:
                ch_num = int(val_match.group(1))
                try:
                    val = float(val_match.group(2))
                    # 如果是 CH0 且 temp_channel_data 已经有数据，说明上一组结束了
                    if ch_num == 0 and temp_channel_data:
                        current_section_data.append(temp_channel_data)
                        temp_channel_data = {}

                    # 初始化或更新当前通道数据
                    temp_channel_data['CH{}'.format(ch_num)] = {'VAL': val, 'ANGLE': None}
                except (ValueError, IndexError):
                    print("  Warning: Could not parse VAL data in line: %s" % line_stripped) # Python 3: 无需 encode
                continue  # 处理完 VAL 后继续下一行

            # 在目标章节内，匹配 Meas ANGLE Check pattern
            angle_match = re.search(
                r'Meas ANGLE Check<<C>> CH(\d+):\s*([-\d.]+)',
                line_u
            )
            if angle_match:
                ch_num = int(angle_match.group(1))
                try:
                    angle = float(angle_match.group(2))
                    # 更新对应通道的 ANGLE 数据
                    ch_key = 'CH{}'.format(ch_num)
                    if ch_key in temp_channel_data: # 修正: 添加了条件判断
                        temp_channel_data[ch_key]['ANGLE'] = angle
                    else:
                        # 如果 ANGLE 先于 VAL 出现（理论上不太可能，但做一下容错）
                        temp_channel_data[ch_key] = {'VAL': None, 'ANGLE': angle}
                except (ValueError, IndexError):
                    print("  Warning: Could not parse ANGLE data in line: %s" % line_stripped) # Python 3: 无需 encode
                continue  # 处理完 ANGLE 后继续下一行

    # 文件读取结束后，处理最后一个目标章节（如果有的话）
    if in_target_section and (current_section_data or temp_channel_data):
        # 保存最后一组未保存的临时数据（如果有的话）
        if temp_channel_data: # 修正: 添加了条件判断
            current_section_data.append(temp_channel_data)
        all_section_data.append({
            'section_name': current_section,
            'data': current_section_data
        })

    # 将结构化的数据转换为扁平化的列表，用于输出
    flattened_data = []
    for section_info in all_section_data: # 修正: 添加了条件判断
        section_name = section_info['section_name']
        section_datasets = section_info['data']

        # 添加章节标题行
        flattened_data.append(
            {'Source_File': os.path.basename(filepath), 'Section': section_name, 'CH_Label': '', 'VAL': '',
             'ANGLE': ''})

        # 添加表头行
        header_row = {'Source_File': '', 'Section': 'CH_Label', 'CH_Label': 'CH_Label', 'VAL': 'VAL', 'ANGLE': 'ANGLE'}
        flattened_data.append(header_row)

        # 添加数据行
        for dataset in section_datasets:
            for ch_label in sorted(dataset.keys(), key=lambda x: int(x[2:])):  # 按 CH 数字排序
                ch_data = dataset[ch_label]
                data_row = {
                    'Source_File': '',
                    'Section': '',
                    'CH_Label': ch_label,
                    'VAL': ch_data.get('VAL', ''),
                    'ANGLE': ch_data.get('ANGLE', '')
                }
                flattened_data.append(data_row)

            # 在每组数据后添加一个空行
            flattened_data.append({'Source_File': '', 'Section': '', 'CH_Label': '', 'VAL': '', 'ANGLE': ''})

    return flattened_data


def write_sheet_to_excel(sheet, data, start_row=1):
    """将数据写入给定的 Excel sheet"""
    if not data: # 修正: 添加了条件判断
        # 如果没有数据，可以写入提示或留空
        sheet.cell(row=start_row, column=1, value="No data found for this section.")
        return start_row + 1

    # 定义列名和对应的列索引
    headers = ['Source_File', 'Section', 'CH_Label', 'VAL', 'ANGLE']

    # 写入数据
    current_row = start_row
    for row_data in data: # 修正: 添加了条件判断
        for col_idx, header in enumerate(headers, 1):
            cell_value = row_data.get(header, '')
            sheet.cell(row=current_row, column=col_idx, value=cell_value)
        current_row += 1

    # 自动调整列宽 (简单处理)
    for col_idx, header in enumerate(headers, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for row_num in range(1, current_row):
            cell_value = sheet.cell(row=row_num, column=col_idx).value
            if cell_value:
                try:
                    cell_len = len(str(cell_value))
                except:
                    cell_len = 0
                if cell_len > max_length:
                    max_length = cell_len
        adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
        sheet.column_dimensions[column].width = adjusted_width

    return current_row  # 返回下一个可用行号


# 修改 get_input 函数以兼容 Python 3
def get_input(prompt):
    """Safe input function for Python 3"""
    # 移除 Python 2 的编码处理部分
    try:
        # Python 3 的 input() 返回字符串，无需 decode
        return input(prompt).strip()
    except (EOFError, KeyboardInterrupt):
        # 处理 Ctrl+C 或意外的 EOF
        print("\nOperation cancelled by user.")
        return 'quit'

def main():
    print("Meas VAL/ANGLE Data Extractor for Specific Sections (Batch Mode - Excel Multi-Sheet - Python 3)")
    print("Processes all .txt files in a given folder.")
    print("Extracts CHx VAL and ANGLE from lines like:")
    print("  Meas VAL Check<<C>> CH2: 57.73115 ...")
    print("  Meas ANGLE Check<<C>> CH2: -0.029 ...")
    print("Only processes data under these sections:")
    for sec in TARGET_SECTIONS:
        print("  - %s" % sec.replace(r'\s+', ' ').replace(r'\.', '.'))
    print("Data is grouped by sections and CH groups, separated by blank lines in output.")
    print("Output is a single .xlsx file with one sheet per .txt file.")
    print("-" * 20)

    while True:
        folder_path = get_input("Enter folder path containing .txt files (or 'quit' to exit): ")

        if folder_path.lower() in ('quit', 'exit', 'q'):
            print("Goodbye!")
            break

        if not folder_path:
            print("Please enter a folder path, or 'quit' to exit.\n")
            continue

        if not os.path.isdir(folder_path):
            print("Error: Path is not a directory or does not exist: %s\n" % folder_path)
            continue

        # 查找文件夹下所有 .txt 文件
        txt_files = [f for f in os.listdir(folder_path) if
                     f.lower().endswith('.txt') and os.path.isfile(os.path.join(folder_path, f))]

        if not txt_files:
            print("No .txt files found in directory: %s\n" % folder_path)
            continue

        print("\nFound %d .txt file(s) in '%s'. Starting processing...\n" % (len(txt_files), folder_path))

        # 创建一个新的 Excel 工作簿
        wb = Workbook()
        # 删除默认创建的 'Sheet'
        default_sheet = wb.active
        wb.remove(default_sheet)

        # 处理每个 .txt 文件
        processed_files_count = 0
        for filename in txt_files:
            filepath = os.path.join(folder_path, filename)

            # Extract data
            file_data = extract_data(filepath)

            # 如果提取到数据或未提取到数据，都为该文件创建一个 sheet
            # 获取不带扩展名的文件名作为 sheet 名
            sheet_name_base = os.path.splitext(filename)[0]
            # 清理 sheet 名以符合 Excel 规范
            sheet_name = sanitize_sheet_name(sheet_name_base)

            # 处理 sheet 名冲突（Excel 不允许同名 sheet）
            original_sheet_name = sheet_name
            counter = 1
            while sheet_name in [s.title for s in wb.worksheets]:
                # 如果名称已存在，则添加后缀
                new_name = "{}_{}".format(original_sheet_name, counter)
                sheet_name = sanitize_sheet_name(new_name)
                counter += 1
                # 防止无限循环（虽然极不可能）
                if counter > 1000:
                    sheet_name = sanitize_sheet_name("File_{}".format(counter))

            # 创建新的 sheet
            ws = wb.create_sheet(title=sheet_name)

            # 写入数据到 sheet
            write_sheet_to_excel(ws, file_data)
            processed_files_count += 1
            if not file_data: # 修正: 更明确的条件判断
                print("  Note: No target data found in %s. Empty sheet created." % filename)
            else:
                print("  Data for %s written to sheet '%s'." % (filename, sheet_name))

        # 生成合并的输出 Excel 文件名 (在指定文件夹内)
        output_file = os.path.join(folder_path, 'ALL_VAL_ANGLE_By_Section.xlsx')

        # 保存 Excel 文件
        try:
            wb.save(output_file)
            print("\nSuccess: All data saved to '%s' with %d sheet(s).\n" % (output_file, processed_files_count))
        except Exception as e:
            print("\nError: Failed to save Excel file '%s'. Reason: %s\n" % (output_file, str(e)))


if __name__ == '__main__':
    main()