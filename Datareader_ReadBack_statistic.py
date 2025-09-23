# -*- coding: utf-8 -*-
# Requires: pip install openpyxl
import re
import os
# 确保安装了兼容 Python 3 的 openpyxl 版本
# pip install openpyxl (推荐) 或 pip install openpyxl==2.6.4 (如果必须用旧版)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# python -m venv myenv
# 需要在虚拟环境下运行 myenv\Scripts\activate

# 定义需要提取数据的章节标题列表
TARGET_SECTIONS = [
    r'8\.1\.3\s+P3Y\s+CTVT\s+Cali\s+10A',
    r'8\.1\.6\s+P3Y\s+CTVT\s+Cali\s+1A',
    r'8\.4\s+50Hz\s+Verification2',
    r'8\.4\s+60Hz\s+Verification1',
    r'8\.4\s+60Hz\s+Verification2'
]

# 编译章节匹配的正则表达式
SECTION_PATTERNS = [re.compile(pattern) for pattern in TARGET_SECTIONS]


def sanitize_sheet_name(name, max_len=31):
    """Excel sheet names have limitations. Sanitize the filename for use as a sheet name."""
    # 移除或替换非法字符: \ / ? * [ ]
    illegal_chars = r'\/?*[]:'
    for char in illegal_chars:
        name = name.replace(char, '_')
    # 限制长度 (Excel 限制为 31 个字符)
    if len(name) > max_len:
        name = name[:max_len]
    # 移除首尾空格
    name = name.strip()
    # 确保名称不为空
    if not name:
        name = "Sheet"
    return name


def extract_data(filepath):
    """Extract Val, Ang, DG, OAng data from file for specific sections."""
    all_section_data = []

    if not os.path.exists(filepath):
        print("  Error: File not found '%s'" % filepath)
        return all_section_data

    print("  Reading file: %s" % os.path.basename(filepath))

    current_section = None
    current_section_data = []
    in_target_section = False

    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line_u = line
            line_stripped = line_u.strip()

            # 检查是否进入新的目标章节
            for i, pattern in enumerate(SECTION_PATTERNS):
                if pattern.search(line_stripped):
                    # 如果之前在目标章节并且有数据，则保存之前的章节数据
                    if in_target_section and current_section_data:
                        all_section_data.append({
                            'section_name': current_section,
                            'data': current_section_data
                        })

                    # 开始新的章节
                    current_section = TARGET_SECTIONS[i].replace(r'\s+', ' ').replace(r'\.', '.')
                    current_section_data = []
                    in_target_section = True
                    break  # 找到匹配就跳出循环

            # 如果不在目标章节，继续读取下一行
            if not in_target_section:
                continue

            # 检查是否进入下一个类似 "数字.数字.数字" 或 "数字.数字" 开头的章节
            if re.match(r'^\d+\.\d+(?:\.\d+)?\s', line_stripped):
                # 保存当前章节数据
                if current_section_data:
                    all_section_data.append({
                        'section_name': current_section,
                        'data': current_section_data
                    })
                    # 重置状态
                    current_section = None
                    current_section_data = []
                    in_target_section = False
                continue

            # 在目标章节内，匹配包含通道数据的行
            # 查找包含 "Readback values" 且后面跟着多个通道数据的行
            if "Readback values" in line_stripped and "ch:" in line_stripped:
                # 提取所有通道数据
                temp_channel_data = {}

                # 使用正则表达式找到所有通道的数据
                channel_matches = re.findall(
                    r'ch:\s*(\d+)\s+Val:\s*([-\d.]+)\s+Ang:\s*([-\d.]+)\s+DG:\s*(\d+)\s+OAng:\s*([-\d.]+)',
                    line_u
                )

                # 将所有通道数据作为一个数据组添加
                for match in channel_matches:
                    try:
                        ch_num = int(match[0])
                        val = float(match[1])
                        ang = float(match[2])
                        dg = int(match[3])
                        oang = float(match[4])

                        temp_channel_data['CH{}'.format(ch_num)] = {
                            'VAL': val,
                            'ANGLE': ang,
                            'DG': dg,
                            'OANG': oang
                        }
                    except (ValueError, IndexError) as e:
                        print("  Warning: Could not parse channel  %s" % str(e))
                        continue

                # 如果找到了通道数据，添加到当前章节数据中
                if temp_channel_data:
                    current_section_data.append(temp_channel_data)

    # 文件读取结束后，处理最后一个目标章节（如果有的话）
    if in_target_section and current_section_data:
        all_section_data.append({
            'section_name': current_section,
            'data': current_section_data
        })

    return all_section_data


def calculate_statistics(section_datasets):
    """计算统计数据"""
    statistics_data = []

    for dataset in section_datasets:
        # 收集数据
        ch0_4_vals = []
        ch5_11_vals = []  # 除了ch8
        ch0_11_vals = []
        ch0_11_angles = []
        ch0_11_oangs = []

        for ch_label in sorted(dataset.keys(), key=lambda x: int(x[2:])):
            ch_num = int(ch_label[2:])
            ch_data = dataset[ch_label]

            # Val数据
            if 0 <= ch_num <= 4:
                ch0_4_vals.append(ch_data['VAL'])
            elif 5 <= ch_num <= 11 and ch_num != 8:
                ch5_11_vals.append(ch_data['VAL'])

            # 所有通道的Val, Angle, OAng数据
            ch0_11_vals.append(ch_data['VAL'])
            ch0_11_angles.append(ch_data['ANGLE'])
            ch0_11_oangs.append(ch_data['OANG'])

        # 计算统计值
        stats = {}

        # Val统计
        if ch0_4_vals:
            stats['VAL_ch0_4_avg'] = sum(ch0_4_vals) / len(ch0_4_vals)
            stats['VAL_ch0_4_max'] = max(ch0_4_vals)
            stats['VAL_ch0_4_min'] = min(ch0_4_vals)
        else:
            stats['VAL_ch0_4_avg'] = stats['VAL_ch0_4_max'] = stats['VAL_ch0_4_min'] = ''

        if ch5_11_vals:
            stats['VAL_ch5_11_avg'] = sum(ch5_11_vals) / len(ch5_11_vals)
            stats['VAL_ch5_11_max'] = max(ch5_11_vals)
            stats['VAL_ch5_11_min'] = min(ch5_11_vals)
        else:
            stats['VAL_ch5_11_avg'] = stats['VAL_ch5_11_max'] = stats['VAL_ch5_11_min'] = ''

        if ch0_11_vals:
            stats['VAL_all_avg'] = sum(ch0_11_vals) / len(ch0_11_vals)
            stats['VAL_all_max'] = max(ch0_11_vals)
            stats['VAL_all_min'] = min(ch0_11_vals)
        else:
            stats['VAL_all_avg'] = stats['VAL_all_max'] = stats['VAL_all_min'] = ''

        # Angle统计
        if ch0_11_angles:
            stats['ANGLE_all_avg'] = sum(ch0_11_angles) / len(ch0_11_angles)
            stats['ANGLE_all_max'] = max(ch0_11_angles)
            stats['ANGLE_all_min'] = min(ch0_11_angles)
        else:
            stats['ANGLE_all_avg'] = stats['ANGLE_all_max'] = stats['ANGLE_all_min'] = ''

        # OAng统计
        if ch0_11_oangs:
            stats['OANG_all_avg'] = sum(ch0_11_oangs) / len(ch0_11_oangs)
            stats['OANG_all_max'] = max(ch0_11_oangs)
            stats['OANG_all_min'] = min(ch0_11_oangs)
        else:
            stats['OANG_all_avg'] = stats['OANG_all_max'] = stats['OANG_all_min'] = ''

        statistics_data.append(stats)

    return statistics_data


def write_sheet_to_excel(sheet, section_info, start_row=1):
    """将数据写入给定的 Excel sheet"""
    section_name = section_info['section_name']
    section_datasets = section_info['data']

    if not section_datasets:
        # 如果没有数据，可以写入提示或留空
        sheet.cell(row=start_row, column=1, value="No data found for this section.")
        return start_row + 1

    # 定义列名和对应的列索引
    headers = ['Source_File', 'Section', 'CH_Label', 'VAL', 'ANGLE', 'DG', 'OANG']
    stat_headers = [
        'Stat_Type', 'VAL_ch0_4_Avg', 'VAL_ch0_4_Max', 'VAL_ch0_4_Min',
        'VAL_ch5_11_Avg', 'VAL_ch5_11_Max', 'VAL_ch5_11_Min',
        'VAL_All_Avg', 'VAL_All_Max', 'VAL_All_Min',
        'ANGLE_All_Avg', 'ANGLE_All_Max', 'ANGLE_All_Min',
        'OANG_All_Avg', 'OANG_All_Max', 'OANG_All_Min'
    ]

    current_row = start_row

    # 写入章节标题
    sheet.cell(row=current_row, column=1, value=os.path.basename(section_info.get('filepath', '')))
    sheet.cell(row=current_row, column=2, value=section_name)
    current_row += 1

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        sheet.cell(row=current_row, column=col_idx, value=header)
    current_row += 1

    # 写入原始数据
    for dataset in section_datasets:
        for ch_label in sorted(dataset.keys(), key=lambda x: int(x[2:])):  # 按 CH 数字排序
            ch_data = dataset[ch_label]
            sheet.cell(row=current_row, column=1, value='')  # Source_File
            sheet.cell(row=current_row, column=2, value='')  # Section
            sheet.cell(row=current_row, column=3, value=ch_label)  # CH_Label
            sheet.cell(row=current_row, column=4, value=ch_data.get('VAL', ''))  # VAL
            sheet.cell(row=current_row, column=5, value=ch_data.get('ANGLE', ''))  # ANGLE
            sheet.cell(row=current_row, column=6, value=ch_data.get('DG', ''))  # DG
            sheet.cell(row=current_row, column=7, value=ch_data.get('OANG', ''))  # OANG
            current_row += 1
        current_row += 2  # 数据组之间空2行

    # 计算统计数据
    statistics_data = calculate_statistics(section_datasets)

    # 添加分隔行
    current_row += 2

    # 写入统计表头
    for col_idx, header in enumerate(stat_headers, 1):
        sheet.cell(row=current_row, column=col_idx, value=header)
    current_row += 1

    # 写入统计数据
    for stats in statistics_data:
        stat_row_data = [
            'Statistics',
            stats.get('VAL_ch0_4_avg', ''),
            stats.get('VAL_ch0_4_max', ''),
            stats.get('VAL_ch0_4_min', ''),
            stats.get('VAL_ch5_11_avg', ''),
            stats.get('VAL_ch5_11_max', ''),
            stats.get('VAL_ch5_11_min', ''),
            stats.get('VAL_all_avg', ''),
            stats.get('VAL_all_max', ''),
            stats.get('VAL_all_min', ''),
            stats.get('ANGLE_all_avg', ''),
            stats.get('ANGLE_all_max', ''),
            stats.get('ANGLE_all_min', ''),
            stats.get('OANG_all_avg', ''),
            stats.get('OANG_all_max', ''),
            stats.get('OANG_all_min', '')
        ]

        for col_idx, value in enumerate(stat_row_data, 1):
            sheet.cell(row=current_row, column=col_idx, value=value)
        current_row += 1
        current_row += 1  # 统计数据之间空1行

    # 自动调整列宽 (简单处理)
    all_headers = headers + stat_headers
    for col_idx in range(1, len(all_headers) + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for row_num in range(start_row, current_row):
            cell_value = sheet.cell(row=row_num, column=col_idx).value
            if cell_value:
                try:
                    cell_len = len(str(cell_value))
                except:
                    cell_len = 0
                if cell_len > max_length:
                    max_length = cell_len
        adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
        sheet.column_dimensions[column].width = adjusted_width

    return current_row  # 返回下一个可用行号


# 修改 get_input 函数以兼容 Python 3
def get_input(prompt):
    """Safe input function for Python 3"""
    # 移除 Python 2 的编码处理部分
    try:
        # Python 3 的 input() 返回字符串，无需 decode
        return input(prompt).strip()
    except (EOFError, KeyboardInterrupt):
        # 处理 Ctrl+C 或意外的 EOF
        print("\nOperation cancelled by user.")
        return 'quit'


def main():
    print("Val/Ang/DG/OAng Data Extractor for Specific Sections (Batch Mode - Excel Multi-Sheet - Python 3)")
    print("Processes all .txt files in a given folder.")
    print("Extracts CHx Val, Ang, DG, OAng from lines containing Readback values.")
    print("Only processes data under these sections:")
    for sec in TARGET_SECTIONS:
        print("  - %s" % sec.replace(r'\s+', ' ').replace(r'\.', '.'))
    print("Data is grouped by sections and CH groups, separated by blank lines in output.")
    print("Output is a single .xlsx file with one sheet per .txt file.")
    print("-" * 20)

    while True:
        folder_path = get_input("Enter folder path containing .txt files (or 'quit' to exit): ")

        if folder_path.lower() in ('quit', 'exit', 'q'):
            print("Goodbye!")
            break

        if not folder_path:
            print("Please enter a folder path, or 'quit' to exit.\n")
            continue

        if not os.path.isdir(folder_path):
            print("Error: Path is not a directory or does not exist: %s\n" % folder_path)
            continue

        # 查找文件夹下所有 .txt 文件
        txt_files = [f for f in os.listdir(folder_path) if
                     f.lower().endswith('.txt') and os.path.isfile(os.path.join(folder_path, f))]

        if not txt_files:
            print("No .txt files found in directory: %s\n" % folder_path)
            continue

        print("\nFound %d .txt file(s) in '%s'. Starting processing...\n" % (len(txt_files), folder_path))

        # 创建一个新的 Excel 工作簿
        wb = Workbook()
        # 删除默认创建的 'Sheet'
        default_sheet = wb.active
        wb.remove(default_sheet)

        # 处理每个 .txt 文件
        processed_files_count = 0
        for filename in txt_files:
            filepath = os.path.join(folder_path, filename)

            # Extract data
            section_data_list = extract_data(filepath)

            # 如果提取到数据或未提取到数据，都为该文件创建一个 sheet
            # 获取不带扩展名的文件名作为 sheet 名
            sheet_name_base = os.path.splitext(filename)[0]
            # 清理 sheet 名以符合 Excel 规范
            sheet_name = sanitize_sheet_name(sheet_name_base)

            # 处理 sheet 名冲突（Excel 不允许同名 sheet）
            original_sheet_name = sheet_name
            counter = 1
            while sheet_name in [s.title for s in wb.worksheets]:
                # 如果名称已存在，则添加后缀
                new_name = "{}_{}".format(original_sheet_name, counter)
                sheet_name = sanitize_sheet_name(new_name)
                counter += 1
                # 防止无限循环（虽然极不可能）
                if counter > 1000:
                    sheet_name = sanitize_sheet_name("File_{}".format(counter))

            # 创建新的 sheet
            ws = wb.create_sheet(title=sheet_name)

            # 为每个section写入数据到 sheet
            current_row = 1
            if not section_data_list:
                ws.cell(row=current_row, column=1, value="No target data found in %s" % filename)
                print("  Note: No target data found in %s. Empty sheet created." % filename)
            else:
                for section_info in section_data_list:
                    section_info['filepath'] = filename  # 添加文件名信息
                    current_row = write_sheet_to_excel(ws, section_info, current_row)
                print("  Data for %s written to sheet '%s'." % (filename, sheet_name))

            processed_files_count += 1

        # 生成合并的输出 Excel 文件名 (在指定文件夹内)
        output_file = os.path.join(folder_path, 'ALL_VAL_ANGLE_By_Section_statistic.xlsx')

        # 保存 Excel 文件
        try:
            wb.save(output_file)
            print("\nSuccess: All data saved to '%s' with %d sheet(s).\n" % (output_file, processed_files_count))
        except Exception as e:
            print("\nError: Failed to save Excel file '%s'. Reason: %s\n" % (output_file, str(e)))


if __name__ == '__main__':
    main()